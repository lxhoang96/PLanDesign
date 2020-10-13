import sys
from PyQt5 import QtCore
from PyQt5 import QtWidgets
from PyQt5 import QtGui
from PyQt5 import uic
from openpyxl import load_workbook
from operator import attrgetter
from database import Database


class Part:
    def __init__(self, name='', line='', priority=0, day=0, number=0, startDate=QtCore.QDate(), endDate=QtCore.QDate(), bo_phan='', to_truong='', sl=0, dateReci=QtCore.QDate()):
        self.name = name
        self.line = line
        self.startdate = startDate
        self.day = day
        self.priority = priority
        self.endDate = endDate
        self.number = number
        self.bo_phan = bo_phan
        self.to_truong = to_truong
        self.sl = sl
        self.dateReci = dateReci

    def store_data(self, data=None, data1=None, hPerDay=0):

        for i, j in enumerate(data):
            divine = j[10]
            if not divine.isdigit():
                data[i][10] = 1
            else:
                data[i][10] = int(divine)

        for i, j in enumerate(data):
            prio = j[11]
            if not prio.isdigit():
                data[i][11] = 1
            else:
                data[i][11] = int(prio)
        for i, j in enumerate(data):
            number = j[6]
            if not number.isdigit():
                data[i][6] = 1
            else:
                data[i][6] = int(number)
        if not data:
            pass

        items = []
        lines = []
        priority = []
        for i in range(len(data)):
            print(data1[i][5])
            hour = int(data[i][6]) / int(data[i][10])
            day = int(hour / hPerDay) + 1
            lines.append(data[i][9])
            priority.append(data[i][11])
            item = Part(name=data[i][1], line=data[i][9], priority=data[i][11], day=day, number=data[i][6], bo_phan=data[i][7], to_truong=data[i][8], sl=data[i][10], dateReci=data1[i][5])
            items.append(item)

        new_priority = list(set(priority))
        new_priority.sort()
        new_lines = list(set(lines))
        new_items = []
        for i in range(len(items)):
            for j in range(len(new_lines)):
                new_items.append([])
                if items[i].line == new_lines[j]:
                    new_items[j].append(items[i])
        new_items = list(filter(None, new_items))

        for each in new_items:
            each.sort(key=attrgetter('priority', 'dateReci'))
            each[0].startDate = each[0].dateReci.addDays(2)
            each[0].endDate = each[0].startDate.addDays(each[0].day-1)
            for i in range(1, len(each)):

                each[i].startDate = each[i - 1].endDate.addDays(1)
                each[i].endDate = each[i].startDate.addDays(each[i].day-1)

        final_items = []
        for item in new_items:
            for each in item:
                final_items.append(each)
        return final_items


class CreateDocuments:
    def __init__(self, name='', orderDate=''):
        self.name = name
        self.orderDate = orderDate

    def closePlan(self, data=None, closeDate=''):
        if not data:
            pass

        wb = load_workbook("form_copy.xlsx")
        sheets = wb.sheetnames
        Sheet1 = wb[sheets[0]]
        Sheet1.cell(row=7, column=1).value = self.name
        Sheet1.cell(row=7, column=8).value = closeDate
        # Set value to table
        rows = len(data)
        for row in range(rows):
            for column in range(1,3):
                Sheet1.cell(row=7+row, column=3+column).value = data[row][column]
            Sheet1.cell(row=7+row, column=2).value = data[row][0]
            Sheet1.cell(row=7+row, column=12).value = data[row][4]
            Sheet1.cell(row=7+row, column=15).value = data[row][3]
            Sheet1.cell(row=7+row, column=17).value = data[row][5]


        # keep only first Sheet
        for s in sheets:
            if s != 'B01':
                sheet_name = wb[s]
                wb.remove(sheet_name)
        return wb

    def embryosPlan(self, data=None):
        if not data:
            pass

        wb = load_workbook("form_copy.xlsx")
        sheets = wb.sheetnames
        Sheet1 = wb[sheets[1]]
        # Set value to table
        rows = len(data)
        for row in range(rows):
            Sheet1.cell(row=6 + row, column=1).value = row + 1
            for column in range(0, 2):
                Sheet1.cell(row=6 + row, column=2 + column).value = data[row][column+1]
            for column in range(5,7):
                Sheet1.cell(row=6 + row, column=column).value = data[row][column]

        # keep only second Sheet
        for s in sheets:
            if s != 'B02':
                sheet_name = wb[s]
                wb.remove(sheet_name)
        return wb

    def inventPlan(self, data=None, dateOfCheck='', closeDate='', closingDay_PXXK=''):
        if not data:
            pass

        wb = load_workbook("form_copy.xlsx")
        sheets = wb.sheetnames
        Sheet1 = wb[sheets[2]]
        # Set value to table
        rows = len(data)
        Sheet1.cell(row=6, column=6).value = dateOfCheck
        Sheet1.cell(row=6, column=7).value = closeDate
        Sheet1.cell(row=6, column=9).value = closingDay_PXXK

        for row in range(rows):
            Sheet1.cell(row=6 + row, column=1).value = row + 1
            Sheet1.cell(row=6 + row, column=2).value = data[row][1]
            Sheet1.cell(row=6 + row, column=3).value = data[row][0]
            Sheet1.cell(row=6 + row, column=4).value = data[row][3]
            Sheet1.cell(row=6 + row, column=5).value = data[row][4]

        # keep only second Sheet
        for s in sheets:
            if s != 'B03':
                sheet_name = wb[s]
                wb.remove(sheet_name)
        return wb

    def embryosPlanPart(self, final_items=None, deliDate=QtCore.QDate()):
        # data1, data2
        if not final_items:
            pass

        db = Database()
        # vat_lieu = db.retrieve_data('LOAI-VAT-TU', 'CHITIET', )
        wb = load_workbook("form_copy.xlsx")
        sheets = wb.sheetnames
        Sheet1 = wb[sheets[4]]
        # Set value to table
        rows = len(final_items)
        for row in range(rows):
            dai = db.retrieve_data('dai', 'CHITIET', 'TEN_CHI_TIET', final_items[row].name)
            new_dai = ", ".join(str(i) for i in dai)
            cao = db.retrieve_data('cao', 'CHITIET', 'TEN_CHI_TIET', final_items[row].name)
            new_cao = ", ".join(str(i) for i in cao)
            day = db.retrieve_data('day', 'CHITIET', 'TEN_CHI_TIET', final_items[row].name)
            new_day = ", ".join(str(i) for i in day)
            don_vi = db.retrieve_data('don_vi', 'CHITIET', 'TEN_CHI_TIET', final_items[row].name)
            new_donVi = ", ".join(str(i) for i in don_vi)
            vat_tu = db.retrieve_data('LOAI_VAT_TU', 'CHITIET', 'TEN_CHI_TIET', final_items[row].name)
            new_vat_tu = ", ".join(str(i) for i in vat_tu)
            Sheet1.cell(row=7 + row, column=2).value = final_items[row].name
            Sheet1.cell(row=7 + row, column=3).value = new_vat_tu
            Sheet1.cell(row=7 + row, column=4).value = new_dai
            Sheet1.cell(row=7 + row, column=5).value = new_cao
            Sheet1.cell(row=7 + row, column=6).value = new_day
            Sheet1.cell(row=7 + row, column=7).value = new_donVi
            Sheet1.cell(row=7 + row, column=8).value = final_items[row].number
            Sheet1.cell(row=7 + row, column=9).value = final_items[row].endDate.toString('dd/MM/yyyy')
            # Sheet1.cell(row=7 + row, column=10).value = final_items[row].endDate.toString('dd/MM/yyyy')
            Sheet1.cell(row=7 + row, column=12).value = final_items[row].to_truong

            # Sheet1.cell(row=7 + row, column=9).value = data[row][4]

        # keep only four Sheet
        for s in sheets:
            if s != 'B05':
                sheet_name = wb[s]
                wb.remove(sheet_name)
        return wb

    def productPlan(self, final_items=None):
        if not final_items:
            pass

        wb = load_workbook("form_copy.xlsx")
        sheets = wb.sheetnames
        Sheet1 = wb[sheets[5]]
        # Set value to table
        rows = len(final_items)
        for row in range(rows):
            Sheet1.cell(row=9 + row, column=1).value = row+1
            Sheet1.cell(row=9 + row, column=2).value = final_items[row].name
            Sheet1.cell(row=9 + row, column=3).value = final_items[row].sl
            Sheet1.cell(row=9 + row, column=4).value = final_items[row].number
            Sheet1.cell(row=9 + row, column=8).value = str(final_items[row].bo_phan)
            Sheet1.cell(row=9 + row, column=9).value = str(final_items[row].to_truong)
            Sheet1.cell(row=9 + row, column=6).value = final_items[row].day
            Sheet1.cell(row=9 + row, column=7).value = final_items[row].startDate.toString('dd/MM/yyyy')

            # Sheet1.cell(row=7 + row, column=9).value = data[row][4]

        # keep only four Sheet
        for s in sheets:
            if s != 'B06':
                sheet_name = wb[s]
                wb.remove(sheet_name)
        return wb

    def productPlanPart(self, final_items=None):

        wb = load_workbook("form_copy.xlsx")
        sheets = wb.sheetnames
        Sheet1 = wb[sheets[6]]
        # Set value to table

        for row in range(len(final_items)):
            Sheet1.cell(row=6 + row, column=1).value = row + 1
            Sheet1.cell(row=6 + row, column=2).value = final_items[row].name
            Sheet1.cell(row=6 + row, column=2 + int(final_items[row].startDate.day())).value = '0'
            Sheet1.cell(row=6 + row, column=2 + int(final_items[row].endDate.day())).value = str(final_items[row].number)


        # keep only seven Sheet
        for s in sheets:
            if s != 'B07':
                sheet_name = wb[s]
                wb.remove(sheet_name)
        return wb

    def save(self, wb, part):
        wb.save(part)


class Four(QtWidgets.QMainWindow):
    def __init__(self, parent=None):
        super(Four, self).__init__(parent)
        uic.loadUi('ui/window4.ui', self)
        self.back.clicked.connect(self.on_back)
        self.save.clicked.connect(self.file_save1)
        self.save_2.clicked.connect(self.file_save2)
        self.save_3.clicked.connect(self.file_save3)
        self.save_4.clicked.connect(self.file_save4)
        self.save_5.clicked.connect(self.file_save5)
        self.save_6.clicked.connect(self.file_save6)

    def on_back(self):
        settings = QtCore.QSettings('myorg', 'myapp')

        settings.remove('wbs')
        theclass = Third(self)
        theclass.restoreSettings()
        Four.hide(self)
        theclass.show()

    def file_save1(self):
        settings = QtCore.QSettings('myorg', 'myapp')

        wb = settings.value('wb1', )

        create = CreateDocuments()
        name = QtWidgets.QFileDialog.getSaveFileName(self, 'Save File')
        if name[0]:
            create.save(wb, name[0])
        else:
            pass

    def file_save2(self):
        settings = QtCore.QSettings('myorg', 'myapp')
        wb = settings.value('wb2', )
        create = CreateDocuments()
        name = QtWidgets.QFileDialog.getSaveFileName(self, 'Save File')
        if name[0]:

            create.save(wb, name[0])
        else:
            pass

    def file_save3(self):
        settings = QtCore.QSettings('myorg', 'myapp')
        wb = settings.value('wb3', )
        create = CreateDocuments()
        name = QtWidgets.QFileDialog.getSaveFileName(self, 'Save File')
        if name[0]:

            create.save(wb, name[0])
        else:
            pass

    def file_save4(self):
        settings = QtCore.QSettings('myorg', 'myapp')
        wb = settings.value('wb4', )
        create = CreateDocuments()
        name = QtWidgets.QFileDialog.getSaveFileName(self, 'Save File')
        if name[0]:

            create.save(wb, name[0])
        else:
            pass

    def file_save5(self):
        settings = QtCore.QSettings('myorg', 'myapp')
        wb = settings.value('wb5', )
        create = CreateDocuments()
        name = QtWidgets.QFileDialog.getSaveFileName(self, 'Save File')
        if name[0]:

            create.save(wb, name[0])
        else:
            pass

    def file_save6(self):
        settings = QtCore.QSettings('myorg', 'myapp')
        wb = settings.value('wb6', )
        create = CreateDocuments()
        name = QtWidgets.QFileDialog.getSaveFileName(self, 'Save File')
        if name[0]:
            create.save(wb, name[0])
        else:
            pass


class Third(QtWidgets.QMainWindow):
    def __init__(self, parent=None):
        super(Third, self).__init__(parent)
        uic.loadUi('ui/window3.ui', self)
        reg_ex = QtCore.QRegExp('^[0-9]{1,2}$')
        validator = QtGui.QRegExpValidator(reg_ex)
        self.supplieNum.setValidator(validator)
        self.confirm.clicked.connect(self.on_supplieChange)
        self.clear.clicked.connect(self.clear_table)
        self.back.clicked.connect(self.on_back)
        self.next.clicked.connect(self.on_next)
        self.actionPlan.triggered.connect(self.on_settingPlan)
        self.actionProduce.triggered.connect(self.on_settingProduce)
        self.create.triggered.connect(self.on_create)
        self.update.triggered.connect(self.on_update)

    def on_settingPlan(self):
        dialog = SettingPlan(self)
        dialog.show()

    def on_settingProduce(self):
        dialog = SettingProduce(self)
        dialog.show()

    def on_supplieLine(self, part, number):
        rows = self.supplieTable.rowCount()
        data = []
        db = Database()
        for i in range(rows):
            it = self.supplieTable.cellWidget(i, 0)
            if it and it.text():
                data.append(it.text())

        for i in range(len(part)):
            supplieItem = str(part[i])
            numberItem = str(number[i])
            print(str(part[i]))
            vat_lieu = db.retrieve_data('LOAI_VAT_TU', 'CHITIET', 'TEN_CHI_TIET', str(part[i]))
            new_vat_lieu = ", ".join(str(i) for i in vat_lieu)
            don_vi = db.retrieve_data('don_vi', 'CHITIET', 'TEN_CHI_TIET', str(part[i]))
            new_don_vi = ", ".join(str(i) for i in don_vi)
            if supplieItem in data:
                pass
            else:
                reg_ex = QtCore.QRegExp('^[0-9]{1,100}$')
                validator = QtGui.QRegExpValidator(reg_ex)
                lineEdit = QtWidgets.QLineEdit(numberItem)
                lineEdit.setValidator(validator)
                rowPosition = self.supplieTable.rowCount()
                self.supplieTable.insertRow(rowPosition)
                self.supplieTable.setCellWidget(rowPosition, 0, QtWidgets.QLineEdit(supplieItem))
                self.supplieTable.setCellWidget(rowPosition, 1, QtWidgets.QLineEdit(new_vat_lieu))
                self.supplieTable.setCellWidget(rowPosition, 2, lineEdit)
                self.supplieTable.setCellWidget(rowPosition, 3, QtWidgets.QLineEdit(new_don_vi))
                self.supplieTable.setCellWidget(rowPosition, 4, QtWidgets.QDateEdit(QtCore.QDate(2020, 1, 1)))
                self.supplieTable.setCellWidget(rowPosition, 5, QtWidgets.QDateEdit(QtCore.QDate(2020, 1, 1)))
                # self.supplieTable.setCellWidget(rowPosition, 6, QtWidgets.QDateEdit(QtCore.QDate(2020, 1, 1)))
                # self.supplieTable.setCellWidget(rowPosition, 7, QtWidgets.QDateEdit(QtCore.QDate(2020, 1, 1)))

    def on_supplieChange(self):

        supplieNum = self.supplieNum.text()
        if supplieNum:
            supplieNum = int(supplieNum)

            for i in range(0, supplieNum):
                reg_ex = QtCore.QRegExp('^[0-9]{1,100}$')
                validator = QtGui.QRegExpValidator(reg_ex)
                lineEdit = QtWidgets.QLineEdit()
                lineEdit.setValidator(validator)
                rowPosition = self.supplieTable.rowCount()
                self.supplieTable.insertRow(rowPosition)
                self.supplieTable.setCellWidget(rowPosition, 0, QtWidgets.QLineEdit())
                self.supplieTable.setCellWidget(rowPosition, 1, QtWidgets.QLineEdit())
                self.supplieTable.setCellWidget(rowPosition, 2, lineEdit)
                self.supplieTable.setCellWidget(rowPosition, 3, QtWidgets.QLineEdit())
                self.supplieTable.setCellWidget(rowPosition, 4, QtWidgets.QDateEdit(QtCore.QDate(2020, 1, 1)))
                self.supplieTable.setCellWidget(rowPosition, 5, QtWidgets.QDateEdit(QtCore.QDate(2020, 1, 1)))
                # self.supplieTable.setCellWidget(rowPosition, 6, QtWidgets.QDateEdit(QtCore.QDate(2020, 1, 1)))
                # self.supplieTable.setCellWidget(rowPosition, 7, QtWidgets.QDateEdit(QtCore.QDate(2020, 1, 1)))
        else:
            pass

    def clear_table(self):
        self.supplieTable.setRowCount(0)

    def saveInstance(self):
        settings = QtCore.QSettings('myorg', 'myapp')

        rows = self.supplieTable.rowCount()

        if rows == 0:
            pass
        else:

            data = []
            for row in range(rows):
                data.append([])
                for column in range(0, 4):
                    it = self.supplieTable.cellWidget(row, column)
                    if it and it.text():
                        pass
                    else:
                        item = QtWidgets.QLineEdit()
                        item.setText("")
                        self.supplieTable.setCellWidget(row, column, item)
                    data[row].append(self.supplieTable.cellWidget(row, column).text())
                for column in range(4, 6):
                    data[row].append(self.supplieTable.cellWidget(row, column).date())

            settings.setValue('bang3', data)

    def restoreSettings(self):
        settings = QtCore.QSettings('myorg', 'myapp')

        data = []
        data = settings.value('bang3', data)
        numrows = len(data)
        if numrows == 0:
            pass
        else:
            for row in range(numrows):
                reg_ex = QtCore.QRegExp('^[0-9]{1,100}$')
                validator = QtGui.QRegExpValidator(reg_ex)
                lineEdit = QtWidgets.QLineEdit(str(data[row][2]))
                lineEdit.setValidator(validator)
                rowPosition = self.supplieTable.rowCount()
                self.supplieTable.insertRow(rowPosition)
                self.supplieTable.setCellWidget(rowPosition, 0, QtWidgets.QLineEdit(str(data[row][0])))
                self.supplieTable.setCellWidget(rowPosition, 1, QtWidgets.QLineEdit(str(data[row][1])))
                self.supplieTable.setCellWidget(rowPosition, 2, lineEdit)
                self.supplieTable.setCellWidget(rowPosition, 3, QtWidgets.QLineEdit(str(data[row][3])))
                self.supplieTable.setCellWidget(rowPosition, 4, QtWidgets.QDateEdit(QtCore.QDate(data[row][4])))
                self.supplieTable.setCellWidget(rowPosition, 5, QtWidgets.QDateEdit(QtCore.QDate(data[row][5])))
                # self.supplieTable.setCellWidget(rowPosition, 6, QtWidgets.QDateEdit(QtCore.QDate(data[row][6])))
                # self.supplieTable.setCellWidget(rowPosition, 7, QtWidgets.QDateEdit(QtCore.QDate(data[row][7])))

    def on_create(self):
        dialog = CreateDB(self)
        dialog.show()

    def on_update(self):
        dialog = UpdateDb(self)
        dialog.show()

    def on_back(self):
        theclass = Second(self)
        theclass.restoreSettings()
        Third.hide(self)
        theclass.show()

    def on_next(self):
        Third.saveInstance(self)
        theclass = Four(self)
        # get data
        settings = QtCore.QSettings('myorg', 'myapp')
        orderDate = settings.value('dateOrder', '')
        deliDate = settings.value('deliDate', '')
        name = settings.value('ten', '')
        data = []
        data = settings.value('bang', data)

        settings1 = QtCore.QSettings('myorg', 'mysetting')
        hour = settings1.value('minWorkh', '')
        hour = int(hour)
        closingDay = settings1.value('closingDay', '')
        closingDay = 0 - int(closingDay)

        create = CreateDocuments(name, orderDate)

        closingDay1 = deliDate.addDays(closingDay)
        data1 = []
        data1 = settings.value('bang2', data1)

        dateOfCheck = settings1.value('dateOfCheck', '')
        dateOfCheck = 0 - int(dateOfCheck)
        dateOfCheck1 = deliDate.addDays(dateOfCheck)
        closingDay_PXXK = settings1.value('closingDay_PXXK', '')
        closingDay_PXXK = 0 - int(closingDay_PXXK)
        closingDay_PXXK1 = deliDate.addDays(closingDay_PXXK)

        data2 = []
        data2 = settings.value('bang3', data2)
        for each in data2:
            print(each)
        part = Part()
        final_items = part.store_data(data1, data2, hour)
        # for each in final_items:
        #     print(each)
        wb1 = create.closePlan(data, closingDay1.toString('dd/MM/yyyy'))
        wb2 = create.embryosPlan(data1)
        wb3 = create.inventPlan(data, dateOfCheck1.toString('dd/MM/yyyy'), closingDay1.toString('dd/MM/yyyy'),
                                closingDay_PXXK1.toString('dd/MM/yyyy'))
        wb5 = create.productPlan(final_items)
        wb4 = create.embryosPlanPart(final_items, deliDate)
        wb6 = create.productPlanPart(final_items)

        settings.setValue('wb1', wb1)
        settings.setValue('wb2', wb2)
        settings.setValue('wb3', wb3)
        settings.setValue('wb4', wb4)
        settings.setValue('wb5', wb5)
        settings.setValue('wb6', wb6)
        Third.hide(self)
        theclass.show()


class Second(QtWidgets.QMainWindow):
    def __init__(self, parent=None):
        super(Second, self).__init__(parent)
        uic.loadUi('ui/window2.ui', self)
        reg_ex = QtCore.QRegExp('^[0-9]{1,100}$')
        validator = QtGui.QRegExpValidator(reg_ex)
        self.lineNum.setValidator(validator)
        self.confirm.clicked.connect(self.on_productChange)
        self.clear.clicked.connect(self.clear_table)
        self.back.clicked.connect(self.on_back)
        self.next.clicked.connect(self.on_next)
        self.actionPlan.triggered.connect(self.on_settingPlan)
        self.actionProduce.triggered.connect(self.on_settingProduce)
        self.cal.clicked.connect(self.cal_amount)
        self.create.triggered.connect(self.on_create)
        self.update.triggered.connect(self.on_update)

    def on_settingPlan(self):
        dialog = SettingPlan(self)
        dialog.show()

    def on_settingProduce(self):
        dialog = SettingProduce(self)
        dialog.show()

    def on_productLine(self, part):

        rows = self.lineTable.rowCount()
        db = Database()
        ids = []
        names = []
        soluongs = []
        for i in range(len(part)):
            id = db.retrieve_data('MA_CHI_TIET', 'CHITIET', str(part[i]), '1')
            listname = db.retrieve_data('TEN_CHI_TIET', 'CHITIET', str(part[i]), '1')
            soluong = db.retrieve_data('so_luong', 'CHITIET', str(part[i]), '1')
            ids.append(id)
            names.append(listname)
            soluongs.append(soluong)

        new_name = [each for name in names for each in name]

        new_id = [each for id in ids for each in id]
        new_soluong = [each for sl in soluongs for each in sl]

        data = []
        for i in range(rows):
            it = self.lineTable.cellWidget(i, 2)
            if it and it.text():
                data.append(it.text())

        for i in range(len(new_id)):
            if new_id[i] in data:
                pass
            else:

                reg_ex1 = QtCore.QRegExp('^[0-9]{1,1}$')
                validator = QtGui.QRegExpValidator(reg_ex1)
                lineEdit1 = QtWidgets.QLineEdit()
                lineEdit1.setValidator(validator)
                lineEdit0 = QtWidgets.QLineEdit(str(new_soluong[i]))
                lineEdit0.setValidator(validator)
                reg_ex = QtCore.QRegExp('^[0-9]{1,10}$')
                validator = QtGui.QRegExpValidator(reg_ex)
                lineEdit = QtWidgets.QLineEdit()
                lineEdit.setValidator(validator)
                lineEdit2 = QtWidgets.QLineEdit()
                lineEdit2.setValidator(validator)
                lineEdit3 = QtWidgets.QLineEdit()
                lineEdit3.setValidator(validator)

                rowPosition = self.lineTable.rowCount()
                self.lineTable.insertRow(rowPosition)

                self.lineTable.setCellWidget(rowPosition, 0, QtWidgets.QLineEdit(str(i+1)))
                self.lineTable.setCellWidget(rowPosition, 1, QtWidgets.QLineEdit(str(new_name[i])))
                self.lineTable.setCellWidget(rowPosition, 2, QtWidgets.QLineEdit(str(new_id[i])))
                self.lineTable.setCellWidget(rowPosition, 3, QtWidgets.QLineEdit())
                self.lineTable.setCellWidget(rowPosition, 4, lineEdit0)
                self.lineTable.setCellWidget(rowPosition, 5, lineEdit2)
                self.lineTable.setCellWidget(rowPosition, 6, lineEdit3)
                self.lineTable.setCellWidget(rowPosition, 7, QtWidgets.QLineEdit())
                self.lineTable.setCellWidget(rowPosition, 8, QtWidgets.QLineEdit())
                self.lineTable.setCellWidget(rowPosition, 9, QtWidgets.QLineEdit())
                self.lineTable.setCellWidget(rowPosition, 10, lineEdit)
                self.lineTable.setCellWidget(rowPosition, 11, lineEdit1)

    def on_productChange(self):
        lineNum = self.lineNum.text()
        if lineNum:
            lineNum = int(lineNum)

            for i in range(0, lineNum):
                reg_ex = QtCore.QRegExp('^[0-9]{1,10}$')
                validator = QtGui.QRegExpValidator(reg_ex)
                lineEdit = QtWidgets.QLineEdit()
                lineEdit.setValidator(validator)
                lineEdit2 = QtWidgets.QLineEdit()
                lineEdit2.setValidator(validator)
                lineEdit3 = QtWidgets.QLineEdit()
                lineEdit3.setValidator(validator)
                reg_ex1 = QtCore.QRegExp('^[0-9]{1,1}$')
                validator = QtGui.QRegExpValidator(reg_ex1)
                lineEdit0 = QtWidgets.QLineEdit()
                lineEdit0.setValidator(validator)
                lineEdit1 = QtWidgets.QLineEdit()
                lineEdit1.setValidator(validator)

                rowPosition = self.lineTable.rowCount()
                self.lineTable.insertRow(rowPosition)

                self.lineTable.setCellWidget(rowPosition, 0, QtWidgets.QLineEdit())
                self.lineTable.setCellWidget(rowPosition, 1, QtWidgets.QLineEdit())
                self.lineTable.setCellWidget(rowPosition, 2, QtWidgets.QLineEdit())
                self.lineTable.setCellWidget(rowPosition, 3, QtWidgets.QLineEdit())
                self.lineTable.setCellWidget(rowPosition, 4, lineEdit0)
                self.lineTable.setCellWidget(rowPosition, 5, lineEdit2)
                self.lineTable.setCellWidget(rowPosition, 6, lineEdit3)
                self.lineTable.setCellWidget(rowPosition, 7, QtWidgets.QLineEdit())
                self.lineTable.setCellWidget(rowPosition, 8, QtWidgets.QLineEdit())
                self.lineTable.setCellWidget(rowPosition, 9, QtWidgets.QLineEdit())
                self.lineTable.setCellWidget(rowPosition, 10, lineEdit)
                self.lineTable.setCellWidget(rowPosition, 11, lineEdit1)
        else:
            pass

    def resize_amount(self):
        settings = QtCore.QSettings('myorg', 'myapp')
        data = []
        data = settings.value('bang', data)
        part = []
        amount = []
        stock = []

        for each in data:
            part.append(each[2])
            amount.append(each[3])
            stock.append(each[6])
        for n, i in enumerate(amount):
            if not i.isdigit():
                amount[n] = 0
            else:
                amount[n] = int(i)
        for n, i in enumerate(stock):
            if not i.isdigit():
                stock[n] = 0
            else:
                stock[n] = int(i)

        db = Database()
        total_number = []
        for i in range(len(part)):
            number = int(amount[i]) - int(stock[i])
            soluong = db.retrieve_data('so_luong', 'CHITIET', str(part[i]), '1')

            numbers = [int(each)*number for each in soluong]
            total_number.append(numbers)

        new_number = [each for number in total_number for each in number]

        return new_number

    def cal_amount(self):
        amount = self.resize_amount()
        rows = self.lineTable.rowCount()
        settings = QtCore.QSettings('myorg', 'mysetting')
        inProgress = settings.value('inProgress', '')
        endProgress = settings.value('endProgress', '')

        for row in range(rows):
            column4 = self.lineTable.cellWidget(row, 4)
            if not column4.text():
                column4.setText('0')
            coef = column4.text()
            column5 = self.lineTable.cellWidget(row, 5)
            if not column5.text():
                column5.setText('0')
            odd = column5.text()

            column3 = self.lineTable.cellWidget(row, 3)
            if column3.text() == "":
                column3.setText('n')
            coef2 = 0
            if str(column3.text()) == 'y':
                coef2 = str(inProgress)
            elif str(column3.text()) == 'n':
                coef2 = str(endProgress)
            need = int((amount[row] * int(coef) - int(odd))*float(coef2))
            if need < 0:
                need = 0
            column6 = self.lineTable.cellWidget(row, 6)
            column6.setText(str(need))

    def clear_table(self):
        self.lineTable.setRowCount(0)

    def on_create(self):
        dialog = CreateDB(self)
        dialog.show()

    def on_update(self):
        dialog = UpdateDb(self)
        dialog.show()

    def saveInstance(self):
        settings = QtCore.QSettings('myorg', 'myapp')

        rows = self.lineTable.rowCount()
        columns = self.lineTable.columnCount()

        data = []
        for row in range(rows):
            data.append([])
            for column in range(columns):
                it = self.lineTable.cellWidget(row, column)
                if it and it.text():
                    pass
                else:
                    item = QtWidgets.QLineEdit()
                    item.setText("")
                    self.lineTable.setCellWidget(row, column, item)
                data[row].append(self.lineTable.cellWidget(row, column).text())
        data = [item for item in data if item[9]]
        for each in data:
            print(each)
        settings.setValue('bang2', data)

    def restoreSettings(self):
        settings = QtCore.QSettings('myorg', 'myapp')

        data = []
        data = settings.value('bang2', data)
        numrows = len(data)
        if numrows == 0:
            pass
        else:
            for row in range(numrows):
                rowPosition = self.lineTable.rowCount()
                self.lineTable.insertRow(rowPosition)
                reg_ex = QtCore.QRegExp('^[0-9]{1,10}$')
                validator = QtGui.QRegExpValidator(reg_ex)
                lineEdit = QtWidgets.QLineEdit(str(data[row][10]))
                lineEdit.setValidator(validator)
                lineEdit2 = QtWidgets.QLineEdit(str(data[row][5]))
                lineEdit2.setValidator(validator)
                lineEdit3 = QtWidgets.QLineEdit(str(data[row][6]))
                lineEdit3.setValidator(validator)
                reg_ex1 = QtCore.QRegExp('^[0-9]{1,1}$')
                validator = QtGui.QRegExpValidator(reg_ex1)
                lineEdit1 = QtWidgets.QLineEdit(str(data[row][11]))
                lineEdit1.setValidator(validator)
                lineEdit0 = QtWidgets.QLineEdit(str(data[row][4]))
                lineEdit0.setValidator(validator)

                self.lineTable.setCellWidget(rowPosition, 0, QtWidgets.QLineEdit(str(data[row][0])))
                self.lineTable.setCellWidget(rowPosition, 1, QtWidgets.QLineEdit(str(data[row][1])))
                self.lineTable.setCellWidget(rowPosition, 2, QtWidgets.QLineEdit(str(data[row][2])))
                self.lineTable.setCellWidget(rowPosition, 3, QtWidgets.QLineEdit(str(data[row][3])))
                self.lineTable.setCellWidget(rowPosition, 4, lineEdit0)
                self.lineTable.setCellWidget(rowPosition, 5, lineEdit2)
                self.lineTable.setCellWidget(rowPosition, 6, lineEdit3)
                self.lineTable.setCellWidget(rowPosition, 7, QtWidgets.QLineEdit(str(data[row][7])))
                self.lineTable.setCellWidget(rowPosition, 8, QtWidgets.QLineEdit(str(data[row][8])))
                self.lineTable.setCellWidget(rowPosition, 9, QtWidgets.QLineEdit(str(data[row][9])))
                self.lineTable.setCellWidget(rowPosition, 10, lineEdit)
                self.lineTable.setCellWidget(rowPosition, 11, lineEdit1)

    def on_back(self):
        Second.saveInstance(self)
        theclass = First(self)
        theclass.restoreSettings()
        Second.hide(self)
        theclass.show()

    def on_next(self):
        Second.saveInstance(self)
        theclass = Third(self)
        settings = QtCore.QSettings('myorg', 'myapp')


        data = []
        data = settings.value('bang2', data)
        part = []
        number = []

        for each in data:
            part.append(each[1])
            number.append(each[6])
        theclass.on_supplieLine(part, number)
        Second.hide(self)
        theclass.show()


class First(QtWidgets.QMainWindow):
    def __init__(self, parent=None):
        super(First, self).__init__(parent)
        uic.loadUi('ui/window1.ui', self)
        reg_ex = QtCore.QRegExp('^[0-9]{1,5}$')
        validator = QtGui.QRegExpValidator(reg_ex)
        self.productNum.setValidator(validator)
        self.confirm.clicked.connect(self.on_productNum)
        self.clear.clicked.connect(self.clear_table)
        self.next.clicked.connect(self.on_next)
        self.actionNew.triggered.connect(self.on_new)
        self.actionQuit_2.triggered.connect(self.close)
        self.actionPlan.triggered.connect(self.on_settingPlan)
        self.actionProduce.triggered.connect(self.on_settingProduce)
        self.create.triggered.connect(self.on_create)
        self.update.triggered.connect(self.on_update)

    def on_productNum(self):
        productNum = self.productNum.text()

        if productNum:
            productNum = int(productNum)

            for i in range(0, productNum):
                rowPosition = self.tableWidget.rowCount()
                self.tableWidget.insertRow(rowPosition)
                lineEdit = QtWidgets.QLineEdit()
                reg_ex = QtCore.QRegExp('^[0-9]{1,10}$')
                validator = QtGui.QRegExpValidator(reg_ex)
                lineEdit.setValidator(validator)
                lineEdit2 = QtWidgets.QLineEdit()
                reg_ex = QtCore.QRegExp('^[0-9]{1,10}$')
                validator = QtGui.QRegExpValidator(reg_ex)
                lineEdit2.setValidator(validator)
                self.tableWidget.setCellWidget(rowPosition, 0, QtWidgets.QLineEdit())
                self.tableWidget.setCellWidget(rowPosition, 1, QtWidgets.QLineEdit())
                self.tableWidget.setCellWidget(rowPosition, 2, QtWidgets.QLineEdit())
                self.tableWidget.setCellWidget(rowPosition, 3, lineEdit)
                self.tableWidget.setCellWidget(rowPosition, 4, lineEdit2)
                self.tableWidget.setCellWidget(rowPosition, 5, QtWidgets.QLineEdit())
                self.tableWidget.setCellWidget(rowPosition, 6, QtWidgets.QLineEdit())
                self.tableWidget.setCellWidget(rowPosition, 7, QtWidgets.QDateEdit(QtCore.QDate(2020, 1, 1)))

        else:
            pass

    def clear_table(self):
        self.tableWidget.setRowCount(0)

    def on_settingPlan(self):
        dialog = SettingPlan(self)
        dialog.show()

    def on_settingProduce(self):
        dialog = SettingProduce(self)
        dialog.show()

    def on_create(self):
        dialog = CreateDB(self)
        dialog.show()

    def on_update(self):
        dialog = UpdateDb(self)
        dialog.show()

    def on_next(self):
        theclass = Second(self)
        First.saveInstance(self)

        settings = QtCore.QSettings('myorg', 'myapp')
        if settings.contains('bang2'):
            theclass.restoreSettings()

        data = []
        data = settings.value('bang', data)
        item = []

        for each in data:

            item.append(each[2])

        theclass.on_productLine(item)
        First.hide(self)
        theclass.show()

    def on_new(self):
        settings = QtCore.QSettings('myorg', 'myapp')
        settings.clear()

        self.hide()
        self.show()

    def saveInstance(self):
        settings = QtCore.QSettings('myorg', 'myapp')
        settings.setValue('ten', self.name.text())
        settings.setValue('loai', self.productNum.text())
        settings.setValue('dateOrder', self.dateOfOrder.date())
        settings.setValue('deliDate', self.deliDate.date())

        rows = self.tableWidget.rowCount()
        columns = self.tableWidget.columnCount()
        data = []
        for row in range(rows):
            data.append([])
            for column in range(columns-1):
                it = self.tableWidget.cellWidget(row, column)
                if it and it.text():
                    pass
                else:
                    item = QtWidgets.QLineEdit()
                    item.setText("")
                    self.tableWidget.setCellWidget(row, column, item)
                data[row].append(self.tableWidget.cellWidget(row, column).text())
            data[row].append(self.tableWidget.cellWidget(row, 7).date())
        settings.setValue('bang', data)

    def restoreSettings(self):
        settings = QtCore.QSettings('myorg', 'myapp')
        name = settings.value('ten', '')
        if name == '':
            pass
        else:
            productNum = settings.value('loai', '')
            orderDate = settings.value('dateOrder', '')
            deliDate = settings.value('deliDate', '')

            self.name.setText(name)
            self.productNum.setText(productNum)
            self.dateOfOrder.setDate(QtCore.QDate(orderDate))
            self.deliDate.setDate(QtCore.QDate(deliDate))

            data = []
            data = settings.value('bang', data)
            numrows = len(data)
            if numrows == 0:
                pass
            else:
                for row in range(numrows):
                    rowPosition = self.tableWidget.rowCount()
                    self.tableWidget.insertRow(rowPosition)
                    lineEdit = QtWidgets.QLineEdit(str(data[row][3]))
                    reg_ex = QtCore.QRegExp('^[0-9]{1,1000}$')
                    validator = QtGui.QRegExpValidator(reg_ex)
                    lineEdit.setValidator(validator)
                    lineEdit2 = QtWidgets.QLineEdit(str(data[row][4]))
                    reg_ex = QtCore.QRegExp('^[0-9]{1,2}$')
                    validator = QtGui.QRegExpValidator(reg_ex)
                    lineEdit2.setValidator(validator)
                    self.tableWidget.setCellWidget(rowPosition, 0, QtWidgets.QLineEdit(str(data[row][0])))
                    self.tableWidget.setCellWidget(rowPosition, 1, QtWidgets.QLineEdit(str(data[row][1])))
                    self.tableWidget.setCellWidget(rowPosition, 2, QtWidgets.QLineEdit(str(data[row][2])))
                    self.tableWidget.setCellWidget(rowPosition, 3, lineEdit)
                    self.tableWidget.setCellWidget(rowPosition, 4, lineEdit2)
                    self.tableWidget.setCellWidget(rowPosition, 5, QtWidgets.QLineEdit(str(data[row][5])))
                    self.tableWidget.setCellWidget(rowPosition, 6, QtWidgets.QLineEdit(str(data[row][6])))
                    self.tableWidget.setCellWidget(rowPosition, 7, QtWidgets.QDateEdit(QtCore.QDate(data[row][7])))


class SettingPlan(QtWidgets.QDialog):
    def __init__(self, parent=None):
        super(SettingPlan, self).__init__(parent)
        uic.loadUi('ui/settingPlan.ui', self)
        self.save.clicked.connect(self.on_save)
        self.restoreSettings()

    def on_save(self):
        settings = QtCore.QSettings('myorg', 'mysetting')
        settings.setValue('dateOfCheck', self.dateOfCheck.text())
        settings.setValue('closingDay', self.closingDay.text())
        settings.setValue('closingDay_PXXK', self.closingDay_PXXK.text())
        settings.setValue('minWorkh', self.minWorkh.text())
        settings.setValue('maxWorkh', self.maxWorkh.text())

    def restoreSettings(self):
        settings = QtCore.QSettings('myorg', 'mysetting')
        dateOfCheck = settings.value('dateOfCheck', '')
        closingDay = settings.value('closingDay', '')
        closingDay_PXXK = settings.value('closingDay_PXXK', '')
        minWorkh = settings.value('minWorkh', '')
        maxWorkh = settings.value('maxWorkh', '')

        self.dateOfCheck.setText(dateOfCheck)
        self.closingDay.setText(closingDay)
        self.closingDay_PXXK.setText(closingDay_PXXK)
        self.minWorkh.setText(minWorkh)
        self.maxWorkh.setText(maxWorkh)


class SettingProduce(QtWidgets.QDialog):
    def __init__(self, parent=None):
        super(SettingProduce, self).__init__(parent)
        uic.loadUi('ui/settingProduce.ui', self)
        self.save.clicked.connect(self.on_save)
        self.restoreSettings()

    def on_save(self):
        settings = QtCore.QSettings('myorg', 'mysetting')
        settings.setValue('inProgress', self.inProgress.text())
        settings.setValue('endProgress', self.endProgress.text())

    def restoreSettings(self):
        settings = QtCore.QSettings('myorg', 'mysetting')
        inProgress = settings.value('inProgress', '')
        endProgress = settings.value('endProgress', '')

        self.inProgress.setText(inProgress)
        self.endProgress.setText(endProgress)


class CreateDB(QtWidgets.QDialog):
    def __init__(self, parent=None):
        super(CreateDB, self).__init__(parent)
        uic.loadUi('ui/create_db.ui', self)
        self.save.clicked.connect(self.on_save)
        self.choose_file.clicked.connect(self.choose)

    def choose(self):
        path = QtWidgets.QFileDialog.getOpenFileName()
        self.path = path

    def on_save(self):
        db = Database()
        name = self.name.text()
        type = str(self.type.currentText())
        if type == 'Part1':
            db.create_parts1_table(name)

        # elif type == 'Part2':
        #     db.create_parts2_table(name)
        # db.add_data(str(self.path), str(name))


class UpdateDb(QtWidgets.QDialog):
    def __init__(self, parent=None):
        super(UpdateDb, self).__init__(parent)
        uic.loadUi('ui/update_db.ui', self)
        self.update.clicked.connect(self.on_update)
        self.choose_file.clicked.connect(self.choose)

    def choose(self):
        path = QtWidgets.QFileDialog.getOpenFileName()
        self.path = path

    def on_update(self):
        db = Database()
        name = self.name.text()
        db.delete_all_tasks(str(name))
        db.add_data(str(self.path), str(name))


def main():
    app = QtWidgets.QApplication(sys.argv)
    db = Database()
    db.create_db()
    # if db.is_table('Items'):
    #     pass
    # else:
    #     db.create_item_table()
    if db.is_table('CHITIET'):
        print('CHITIET has been created.')
    else:
        db.create_parts1_table('CHITIET')
        db.add_data('example_db/file_thu.csv', 'CHITIET')
    # if db.is_table('Part2'):
    #     pass
    # else:
    #     db.create_parts2_table('Part2')
    #     db.add_data('example_db/bep11524_part2.csv', 'Part2')

    settings = QtCore.QSettings('myorg', 'myapp')
    settings.remove('bang2')
    settings.remove('bang3')

    home = First()
    # main.restoreSettings()
    home.show()

    sys.exit(app.exec_())


if __name__ == '__main__':
    main()
